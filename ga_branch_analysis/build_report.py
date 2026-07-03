"""GA 지사분석 보고서 엑셀 파일 생성 (메인 오케스트레이션)."""
import os
import shutil

from openpyxl import Workbook

import scoring as sc
import clustering as cl
import preprocessing_qna as pq
import charts as ch
import report_style as rs

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_PATH = os.path.join(BASE_DIR, "data", "raw_data.csv")
CHART_DIR = os.path.join(BASE_DIR, "charts_tmp")
OUTPUT_PATH = os.path.join(BASE_DIR, "output", "GA지사분석_보고서.xlsx")

REPORT_DATE = "2026-07-03"

FLOAT_FMT = "#,##0.00"
INT_FMT = "0"
PCT_FMT = "0.0"


def prepare_data():
    df1 = sc.load_group1(DATA_PATH)
    df1 = sc.add_kpi_scores(df1)
    df1 = sc.add_category_pct(df1)
    df1 = cl.assign_primary_tier(df1)
    df1, meta = cl.refine_middle_tier(df1)
    profile = cl.build_group_profile(df1)
    score_ref = sc.build_score_reference_table(df1)
    full = sc.load_raw(DATA_PATH)
    return df1, meta, profile, score_ref, full


def generate_charts(df1, profile):
    os.makedirs(CHART_DIR, exist_ok=True)
    paths = {}

    paths["group_counts"] = os.path.join(CHART_DIR, "01_group_counts.png")
    ch.chart_group_counts(df1, paths["group_counts"])

    paths["score_box"] = os.path.join(CHART_DIR, "02_score_box.png")
    ch.chart_score_boxplot(df1, paths["score_box"])

    cat_cols = [c for c in profile.columns if c.endswith("_평균달성률")]
    paths["radar"] = os.path.join(CHART_DIR, "03_radar.png")
    ch.chart_category_radar(profile, cat_cols, paths["radar"])

    dist_targets = [
        ("보장성NCEV_가상_점수", "보장성NCEV 점수 분포 (배점 20)"),
        ("장기1차년도손해율_가상_점수", "장기1차년도손해율 점수 분포 (배점 10)"),
        ("유지율_4_가상_점수", "유지율_4 점수 분포 (배점 10)"),
        ("가동5만_가상_점수", "가동5만 점수 분포 (배점 5)"),
    ]
    paths["dist"] = []
    for i, (col, title) in enumerate(dist_targets):
        p = os.path.join(CHART_DIR, f"04_dist_{i}.png")
        ch.chart_score_distribution(df1, col, p, title)
        paths["dist"].append(p)

    return paths


def generate_qna_charts(full):
    q4 = pq.q4_group_dept_ncev(full)
    q5 = pq.q5_group_dept_utilization(full)
    p4 = os.path.join(CHART_DIR, "q4_hist.png")
    p5 = os.path.join(CHART_DIR, "q5_box.png")
    ch.chart_q4_hist(q4, p4)
    ch.chart_q5_box(q5, p5)
    return q4, q5, p4, p5


def build_sheet_summary(wb, df1, profile, chart_paths):
    ws = wb.active
    ws.title = "00_요약"
    rs.sheet_title(ws, "GA 지사분석 보고서 — 1그룹 지사 성과 그룹화")
    rs.note(ws, "A2", f"분석대상: 그룹분류=1그룹 지사 {len(df1)}개 (사업단 {df1['사업단_가상'].nunique()}개, 조직 {df1['조직명_가상'].nunique()}개) | 작성일: {REPORT_DATE}")
    rs.note(ws, "A3", "방법론: 11개 성과지표를 지표별 배점(합계 100점) 내에서 분위수(rank 기반) 등분할로 채점 → 각 점수대 지사 수를 균등 배분 → 종합점수 기준 KMeans(k=3)로 우수형/중간형/부진형 분류 → 중간형은 실루엣 스코어로 최적 k를 자동 선택해 세분화")

    ws["A5"] = "그룹유형별 요약"
    ws["A5"].font = rs.SECTION_FONT
    show_cols = ["그룹유형", "지사수", "평균종합점수", "그룹특징"]
    next_row = rs.write_df(ws, profile[show_cols].round(1), start_row=6,
                            number_formats={"평균종합점수": FLOAT_FMT})

    ws[f"A{next_row + 2}"] = "핵심 인사이트"
    ws[f"A{next_row + 2}"].font = rs.SECTION_FONT
    insights = [
        f"- 1그룹 {len(df1)}개 지사는 우수형/중간형(2개 세부유형)/부진형 4개 유형으로 분류됨.",
        f"- 우수형은 전 지표에서 고르게 상위권이며 특히 수익성·활동성 지표가 강함.",
        f"- 중간형은 세부적으로 '수익성 우수형'과 '손해율 강점·수익성 약점형'으로 나뉘어, 동일한 '중간' 총점이라도 강약점 프로파일이 다름.",
        f"- 부진형은 손해율 카테고리 달성률이 가장 낮아 손해율 관리가 개선 포인트로 파악됨.",
    ]
    for i, line in enumerate(insights):
        ws[f"A{next_row + 3 + i}"] = line
        ws[f"A{next_row + 3 + i}"].font = rs.BODY_FONT

    img_row = next_row + 3 + len(insights) + 2
    rs.insert_image(ws, chart_paths["group_counts"], f"A{img_row}", width=430, height=260)
    rs.insert_image(ws, chart_paths["score_box"], f"H{img_row}", width=460, height=260)

    rs.autofit_columns(ws)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = rs.NAVY
    return ws


def build_sheet_score_ref(wb, score_ref):
    ws = wb.create_sheet("01_점수산출기준")
    rs.sheet_title(ws, "지표별 점수산출 기준 (분위수 등분할)")
    rs.note(ws, "A2", "각 지표를 배점 구간 수만큼 rank 기반으로 등분해 점수를 부여 (점수대별 지사 수가 최대한 균등). 손해율 지표는 방향을 반전(낮을수록 고득점).")
    rs.write_df(ws, score_ref, start_row=4,
                number_formats={"원자료_최소": FLOAT_FMT, "원자료_최대": FLOAT_FMT})
    rs.autofit_columns(ws)
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = rs.NAVY
    return ws


def build_sheet_branch_data(wb, df1):
    ws = wb.create_sheet("02_지사별_점수데이터")
    rs.sheet_title(ws, "지사별 원자료 및 점수 데이터 (1그룹)")

    display = df1.sort_values("종합점수", ascending=False).reset_index(drop=True)
    key_cols = ["상호명_가상", "사업단_가상", "조직명_가상"]
    kpi_cols = sc.KPI_COLS
    score_cols = sc.SCORE_COLS
    tail_cols = ["종합점수", "그룹유형", "그룹특징"]
    ordered = key_cols + kpi_cols + score_cols + tail_cols
    display = display[ordered]

    num_fmt = {c: FLOAT_FMT for c in kpi_cols}
    num_fmt.update({c: INT_FMT for c in score_cols})
    num_fmt["종합점수"] = INT_FMT

    rs.write_df(ws, display, start_row=3, number_formats=num_fmt)
    rs.autofit_columns(ws, max_width=22)
    ws.freeze_panes = "D4"

    score_col_letter = None
    from openpyxl.utils import get_column_letter
    score_col_idx = ordered.index("종합점수") + 1
    score_col_letter = get_column_letter(score_col_idx)
    rs.add_color_scale(ws, f"{score_col_letter}4:{score_col_letter}{3 + len(display)}")

    ws.sheet_properties.tabColor = rs.NAVY
    return ws


def build_sheet_profile(wb, profile, meta, chart_paths):
    ws = wb.create_sheet("03_그룹분류_프로파일")
    rs.sheet_title(ws, "그룹유형별 프로파일")
    rs.note(ws, "A2", f"중간형 세분화: 실루엣 스코어 기준 최적 k={meta['선택된_k']} 자동 선택 (실루엣 스코어={meta['실루엣스코어']})")

    cat_cols = [c for c in profile.columns if c.endswith("_평균달성률")]
    show_cols = ["그룹유형", "지사수", "평균종합점수"] + cat_cols + ["그룹특징"]
    fmt = {c: FLOAT_FMT for c in cat_cols}
    fmt["평균종합점수"] = FLOAT_FMT
    next_row = rs.write_df(ws, profile[show_cols].round(2), start_row=4, number_formats=fmt)

    rs.insert_image(ws, chart_paths["radar"], f"A{next_row + 2}", width=480, height=480)
    rs.autofit_columns(ws)
    ws.sheet_properties.tabColor = rs.NAVY
    return ws


def build_sheet_viz(wb, chart_paths):
    ws = wb.create_sheet("04_시각화")
    rs.sheet_title(ws, "지표별 점수 균등분포 검증")
    rs.note(ws, "A2", "분위수 등분할 방식 적용 결과, 지표별 배점 구간(1~배점)에 지사가 고르게 분포됨을 확인.")

    row = 4
    col_positions = ["A", "L"]
    for i, p in enumerate(chart_paths["dist"]):
        anchor = f"{col_positions[i % 2]}{row}"
        rs.insert_image(ws, p, anchor, width=380, height=230)
        if i % 2 == 1:
            row += 15
    ws.sheet_properties.tabColor = rs.NAVY
    return ws


def build_sheet_qna_1_3(wb, full):
    ws = wb.create_sheet("05_전처리_Q1_Q3")
    rs.sheet_title(ws, "전처리 Q1~Q3 (전체 데이터: 1/2/3그룹)")

    ws["A2"] = "[Q1] 사업단별·조직별 상호명 개수"
    ws["A2"].font = rs.SECTION_FONT
    q1 = pq.q1_count_by_dept_org(full)
    r = rs.write_df(ws, q1, start_row=3, number_formats={"상호명_개수": INT_FMT})

    r += 2
    ws[f"A{r}"] = "[Q2] 사업단별 성과지표 평균"
    ws[f"A{r}"].font = rs.SECTION_FONT
    q2 = pq.q2_dept_means(full)
    fmt2 = {c: FLOAT_FMT for c in pq.Q2_COLS}
    r = rs.write_df(ws, q2.round(2), start_row=r + 1, number_formats=fmt2)

    ranked, summary, top10 = pq.q3_ncev_compare(full, q2)
    r += 2
    ws[f"A{r}"] = "[Q3] 사업단별 평균 보장성NCEV 비교"
    ws[f"A{r}"].font = rs.SECTION_FONT
    ans = (
        f"최고: {summary['최고_사업단']} (평균 {summary['최고_평균NCEV']:,.1f}) / "
        f"최저: {summary['최저_사업단']} (평균 {summary['최저_평균NCEV']:,.1f}) / "
        f"차이: {summary['차이']:,.1f}"
    )
    ws[f"A{r + 1}"] = ans
    ws[f"A{r + 1}"].font = rs.BODY_FONT
    r = rs.write_df(ws, ranked.round(2), start_row=r + 3,
                     number_formats={"보장성NCEV_가상": FLOAT_FMT})

    r += 2
    ws[f"A{r}"] = f"[3-1] {summary['최고_사업단']} 소속 조직 수: {summary['최고사업단_조직수']}개 / 상호명 수: {summary['최고사업단_상호명수']}개"
    ws[f"A{r}"].font = rs.BODY_FONT_BOLD

    r += 2
    ws[f"A{r}"] = f"[3-2] {summary['최고_사업단']} 내 보장성NCEV TOP10 상호명"
    ws[f"A{r}"].font = rs.SECTION_FONT
    rs.write_df(ws, top10.round(2), start_row=r + 1, number_formats={"보장성NCEV_가상": FLOAT_FMT})

    rs.autofit_columns(ws)
    ws.sheet_properties.tabColor = "548235"
    return ws


def build_sheet_qna_4_6(wb, full, q4, q5, p4, p5):
    ws = wb.create_sheet("06_전처리_Q4_Q6")
    rs.sheet_title(ws, "전처리 Q4~Q6 (전체 데이터: 1/2/3그룹)")

    ws["A2"] = "[Q4] 그룹별 사업단 평균 보장성NCEV 분포 (histogram)"
    ws["A2"].font = rs.SECTION_FONT
    rs.insert_image(ws, p4, "A3", width=760, height=230)
    r = rs.write_df(ws, q4.round(2), start_row=17, number_formats={"평균보장성NCEV": FLOAT_FMT})

    r += 2
    ws[f"A{r}"] = "[Q5] 그룹별 사업단 평균 가동5만/가동20만 분포 (boxplot)"
    ws[f"A{r}"].font = rs.SECTION_FONT
    img_row = r + 1
    rs.insert_image(ws, p5, f"A{img_row}", width=640, height=260)
    r = rs.write_df(ws, q5.round(2), start_row=img_row + 15,
                     number_formats={"가동5만_가상": FLOAT_FMT, "가동20만_가상": FLOAT_FMT})

    dept_lr, min_dept, top10_min, top10_max = pq.q6_loss_ratio_increase(full)
    r += 2
    ws[f"A{r}"] = "[Q6] 사업단별 장기1차→2차년도손해율 증가폭"
    ws[f"A{r}"].font = rs.SECTION_FONT
    ws[f"A{r + 1}"] = f"증가폭이 가장 적은 사업단: {min_dept}"
    ws[f"A{r + 1}"].font = rs.BODY_FONT_BOLD
    fmt_lr = {c: FLOAT_FMT for c in ["장기1차년도손해율_가상", "장기2차년도손해율_가상", "증가폭"]}
    r = rs.write_df(ws, dept_lr.round(2), start_row=r + 3, number_formats=fmt_lr)

    r += 2
    ws[f"A{r}"] = f"[6-1] {min_dept} 내 증가폭 최소 TOP10 상호명"
    ws[f"A{r}"].font = rs.SECTION_FONT
    r = rs.write_df(ws, top10_min.round(2), start_row=r + 1, number_formats=fmt_lr)

    r += 2
    ws[f"A{r}"] = f"[6-2] {min_dept} 내 증가폭 최대 TOP10 상호명"
    ws[f"A{r}"].font = rs.SECTION_FONT
    rs.write_df(ws, top10_max.round(2), start_row=r + 1, number_formats=fmt_lr)

    rs.autofit_columns(ws)
    ws.sheet_properties.tabColor = "548235"
    return ws


def main():
    df1, meta, profile, score_ref, full = prepare_data()
    chart_paths = generate_charts(df1, profile)
    q4, q5, p4, p5 = generate_qna_charts(full)
    chart_paths["q4"] = p4
    chart_paths["q5"] = p5

    wb = Workbook()
    build_sheet_summary(wb, df1, profile, chart_paths)
    build_sheet_score_ref(wb, score_ref)
    build_sheet_branch_data(wb, df1)
    build_sheet_profile(wb, profile, meta, chart_paths)
    build_sheet_viz(wb, chart_paths)
    build_sheet_qna_1_3(wb, full)
    build_sheet_qna_4_6(wb, full, q4, q5, p4, p5)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
