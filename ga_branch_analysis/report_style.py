"""엑셀 보고서 공통 서식 헬퍼."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.drawing.image import Image as XLImage

NAVY = "1F3864"
LIGHT_NAVY = "D9E2F3"
WHITE = "FFFFFF"
INK = "262626"

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(color=WHITE, bold=True, size=10, name="맑은 고딕")
TITLE_FONT = Font(bold=True, size=15, color=NAVY, name="맑은 고딕")
SUBTITLE_FONT = Font(size=10, italic=True, color="595959", name="맑은 고딕")
SECTION_FONT = Font(bold=True, size=12, color=NAVY, name="맑은 고딕")
BODY_FONT = Font(size=10, name="맑은 고딕")
BODY_FONT_BOLD = Font(size=10, bold=True, name="맑은 고딕")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")


def sheet_title(ws, text, cell="A1", span=1):
    ws[cell] = text
    ws[cell].font = TITLE_FONT
    if span > 1:
        col = cell[0]
        row = cell[1:]
        end_col = get_column_letter(get_column_letter_index(col) + span - 1)
        ws.merge_cells(f"{cell}:{end_col}{row}")


def get_column_letter_index(letter):
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(letter)


def write_df(ws, df, start_row, start_col=1, number_formats=None, header_wrap=True):
    """DataFrame을 지정 위치에 표 형태(헤더 스타일+테두리)로 기록. 다음 빈 행 인덱스 반환."""
    number_formats = number_formats or {}
    ncols = len(df.columns)

    for j, col_name in enumerate(df.columns):
        c = ws.cell(row=start_row, column=start_col + j, value=str(col_name))
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER

    for i, (_, row) in enumerate(df.iterrows()):
        r = start_row + 1 + i
        for j, col_name in enumerate(df.columns):
            val = row[col_name]
            c = ws.cell(row=r, column=start_col + j, value=val)
            c.font = BODY_FONT
            c.border = BORDER
            c.alignment = LEFT if j == 0 else CENTER
            if col_name in number_formats:
                c.number_format = number_formats[col_name]
            if i % 2 == 1:
                c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    return start_row + 1 + len(df)


def autofit_columns(ws, min_width=9, max_width=42):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            col = cell.column_letter
            widths[col] = max(widths.get(col, 0), length)
    for col, w in widths.items():
        ws.column_dimensions[col].width = min(max(w + 2, min_width), max_width)


def add_color_scale(ws, cell_range):
    rule = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    )
    ws.conditional_formatting.add(cell_range, rule)


def insert_image(ws, path, anchor, width=None, height=None):
    img = XLImage(path)
    if width:
        img.width = width
    if height:
        img.height = height
    ws.add_image(img, anchor)


def note(ws, cell, text):
    ws[cell] = text
    ws[cell].font = SUBTITLE_FONT
    ws[cell].alignment = LEFT
